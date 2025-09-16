"""Safe I/O operations for the Data Wrangler app."""

import pandas as pd
import io
import tempfile
import os
from typing import Dict, List, Optional, Tuple, Union, BinaryIO
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def safe_read_csv(file_path: Union[str, Path, BinaryIO], 
                 encoding: Optional[str] = None,
                 delimiter: str = ',',
                 decimal: str = '.',
                 thousands: Optional[str] = None,
                 sample_size: int = 10000) -> Tuple[pd.DataFrame, str]:
    """Safely read a CSV file with error handling and encoding detection.
    
    Args:
        file_path: Path to CSV file or file-like object
        encoding: File encoding (auto-detected if None)
        delimiter: CSV delimiter
        decimal: Decimal separator
        thousands: Thousands separator
        sample_size: Size of sample for encoding detection
        
    Returns:
        Tuple of (DataFrame, detected_encoding)
    """
    try:
        # Handle file-like objects
        if hasattr(file_path, 'read'):
            file_obj = file_path
            file_name = getattr(file_path, 'name', 'unknown')
        else:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            file_name = str(file_path)
            file_obj = open(file_path, 'rb')
        
        # Detect encoding if not provided
        if encoding is None:
            encoding = detect_encoding_from_file(file_obj, sample_size)
            file_obj.seek(0)  # Reset file pointer
        
        # Read CSV with specified parameters
        df = pd.read_csv(
            file_obj,
            encoding=encoding,
            delimiter=delimiter,
            decimal=decimal,
            thousands=thousands,
            low_memory=False
        )
        
        # Close file if we opened it
        if not hasattr(file_path, 'read'):
            file_obj.close()
        
        logger.info(f"Successfully read {file_name} with encoding {encoding}")
        return df, encoding
        
    except Exception as e:
        logger.error(f"Error reading CSV file {file_name}: {e}")
        raise


def detect_encoding_from_file(file_obj: BinaryIO, sample_size: int = 10000) -> str:
    """Detect encoding from a file object.
    
    Args:
        file_obj: Binary file object
        sample_size: Number of bytes to sample
        
    Returns:
        Detected encoding string
    """
    try:
        import chardet
        
        # Save current position
        current_pos = file_obj.tell()
        
        # Read sample
        sample = file_obj.read(sample_size)
        file_obj.seek(current_pos)  # Restore position
        
        # Detect encoding
        result = chardet.detect(sample)
        encoding = result.get('encoding', 'utf-8')
        confidence = result.get('confidence', 0)
        
        # Fallback to utf-8 if confidence is too low
        if confidence < 0.7:
            logger.warning(f"Low encoding confidence ({confidence:.2f}), using utf-8")
            return 'utf-8'
        
        return encoding
        
    except ImportError:
        logger.warning("chardet not available, using utf-8")
        return 'utf-8'
    except Exception as e:
        logger.warning(f"Error detecting encoding: {e}, using utf-8")
        return 'utf-8'


def safe_write_excel(dataframes: Dict[str, pd.DataFrame], 
                    output_path: Union[str, Path],
                    engine: str = 'openpyxl',
                    format_tables: bool = True) -> bool:
    """Safely write DataFrames to Excel file with optional formatting.
    
    Args:
        dataframes: Dictionary mapping sheet names to DataFrames
        output_path: Path to output Excel file
        engine: Excel engine to use
        format_tables: Whether to apply table formatting and colors
        
    Returns:
        True if successful, False otherwise
    """
    try:
        output_path = Path(output_path)
        
        # Create parent directories if they don't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Write to Excel
        with pd.ExcelWriter(output_path, engine=engine) as writer:
            for sheet_name, df in dataframes.items():
                # Clean sheet name (Excel has restrictions)
                clean_sheet_name = clean_excel_sheet_name(sheet_name)
                df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                
                # Apply formatting if requested
                if format_tables:
                    apply_excel_formatting(writer, clean_sheet_name, df)
        
        logger.info(f"Successfully wrote Excel file: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error writing Excel file {output_path}: {e}")
        return False


def apply_excel_formatting(writer, sheet_name: str, df: pd.DataFrame):
    """Apply formatting to Excel sheet.
    
    Args:
        writer: Excel writer object
        sheet_name: Name of the sheet to format
        df: DataFrame being written
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo
        
        # Get the worksheet
        worksheet = writer.sheets[sheet_name]
        
        # Define colors
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Alternate row colors for better readability
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create a table (Excel table format)
        if len(df) > 0:
            table_range = f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}"
            table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_range)
            
            # Apply table style
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            worksheet.add_table(table)
        
        logger.debug(f"Applied formatting to sheet: {sheet_name}")
        
    except ImportError:
        logger.warning("openpyxl styling not available, skipping formatting")
    except Exception as e:
        logger.warning(f"Could not apply Excel formatting: {e}")


def clean_excel_sheet_name(name: str, max_length: int = 31) -> str:
    """Clean sheet name for Excel compatibility.
    
    Args:
        name: Original sheet name
        max_length: Maximum length allowed by Excel
        
    Returns:
        Cleaned sheet name
    """
    # Remove invalid characters
    invalid_chars = ['[', ']', '*', '?', ':', '\\', '/']
    clean_name = name
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    
    # Truncate if too long
    if len(clean_name) > max_length:
        clean_name = clean_name[:max_length]
    
    # Ensure it's not empty
    if not clean_name:
        clean_name = "Sheet1"
    
    return clean_name


def get_file_info(file_path: Union[str, Path]) -> Dict[str, any]:
    """Get basic file information.
    
    Args:
        file_path: Path to file
        
    Returns:
        Dictionary with file information
    """
    try:
        file_path = Path(file_path)
        stat = file_path.stat()
        
        return {
            'name': file_path.name,
            'size_bytes': stat.st_size,
            'size_mb': stat.st_size / (1024 * 1024),
            'modified': stat.st_mtime,
            'exists': file_path.exists()
        }
    except Exception as e:
        logger.error(f"Error getting file info for {file_path}: {e}")
        return {
            'name': str(file_path),
            'size_bytes': 0,
            'size_mb': 0,
            'modified': 0,
            'exists': False
        }


def validate_file_extension(file_path: Union[str, Path], 
                          allowed_extensions: List[str] = None) -> bool:
    """Validate file extension.
    
    Args:
        file_path: Path to file
        allowed_extensions: List of allowed extensions (default: ['.csv'])
        
    Returns:
        True if extension is allowed
    """
    if allowed_extensions is None:
        allowed_extensions = ['.csv']
    
    file_path = Path(file_path)
    extension = file_path.suffix.lower()
    
    return extension in allowed_extensions


def create_temp_file(suffix: str = '.csv') -> Tuple[str, str]:
    """Create a temporary file.
    
    Args:
        suffix: File suffix
        
    Returns:
        Tuple of (file_path, file_name)
    """
    temp_fd, temp_path = tempfile.mkstemp(suffix=suffix)
    os.close(temp_fd)  # Close the file descriptor
    
    file_name = Path(temp_path).name
    return temp_path, file_name


def cleanup_temp_files(file_paths: List[str]) -> None:
    """Clean up temporary files.
    
    Args:
        file_paths: List of file paths to clean up
    """
    for file_path in file_paths:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.debug(f"Cleaned up temp file: {file_path}")
        except Exception as e:
            logger.warning(f"Could not clean up temp file {file_path}: {e}")


def read_multiple_csvs(file_paths: List[Union[str, Path]], 
                      encoding: Optional[str] = None,
                      delimiter: str = ',',
                      decimal: str = '.',
                      thousands: Optional[str] = None) -> Tuple[Dict[str, pd.DataFrame], Dict[str, str]]:
    """Read multiple CSV files safely.
    
    Args:
        file_paths: List of file paths
        encoding: File encoding (auto-detected if None)
        delimiter: CSV delimiter
        decimal: Decimal separator
        thousands: Thousands separator
        
    Returns:
        Tuple of (dataframes_dict, encodings_dict)
    """
    dataframes = {}
    encodings = {}
    
    for file_path in file_paths:
        try:
            file_name = Path(file_path).stem
            df, detected_encoding = safe_read_csv(
                file_path, encoding, delimiter, decimal, thousands
            )
            dataframes[file_name] = df
            encodings[file_name] = detected_encoding
            
        except Exception as e:
            logger.error(f"Failed to read {file_path}: {e}")
            # Continue with other files
            continue
    
    return dataframes, encodings


def save_report_to_file(report_text: str, output_path: Union[str, Path]) -> bool:
    """Save report text to file.
    
    Args:
        report_text: Report text content
        output_path: Path to output file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report_text)
        
        logger.info(f"Successfully saved report to: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error saving report to {output_path}: {e}")
        return False
